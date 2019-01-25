from flask import Flask, request, session, g, redirect, url_for, render_template,flash
from openpyxl import load_workbook
import os
import sqlite3

app = Flask(__name__)
# session - переменная в которой хранится информация о прошлых действиях пользователя
# e = create_engine("sqlite:///table.db")
pathdb = os.path.join(os.path.dirname(__file__), 'table.db')
app.config.update(DATABASE=pathdb, DEBUG=True, SECRET_KEY='secretkey',
                  USERNAME='admin', PASSWORD='admin')
if os.path.exists(pathdb):
    os.remove(pathdb)


def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(pathdb)
    return db


ws = None
l = None
systemtables = ['sqlite_sequence', 'tablerel']


def gettype(val):
    if val.isdigit():
        return "integer"
    elif val.replace('.', '', 1).isdigit():
        return "double"
    else:
        return "text"


def make_dicts(cursor, row):
    return dict((cursor.description[idx][0], value)
                for idx, value in enumerate(row))


def getPk(tab):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("pragma table_info ('" + tab + "') ")
    data = cursor.fetchall()
    ind1 = []
    for i in data:
        if i[5] == 1:
            return i
def checkandcorrect(st,tab):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("pragma table_info ('" + tab + "') ")
    data = cursor.fetchall()
    ok=True
    while ok:
        ok=False
        for i in data:
            if i[1] == st:
                ok=True
                st+='1'
                break
    return st
@app.route('/Author', methods=['POST', 'GET'])
def author():
    return render_template("Author.html")
@app.route('/reltables', methods=['POST', 'GET'])
def reltables():
    f = open("static\script.txt", 'a')
    db = get_db()
    cursor = db.cursor()
    cursor.execute("select name from sqlite_master where type='table'")
    tab = cursor.fetchall()
    t = []
    nt = '-'
    st = ''
    for i in tab:
        if i[0] not in systemtables:
            t.append(i[0])
    if request.method == 'POST':
        try:
            tab1 = request.form['table1']
            rel = request.form['rel']
            tab2 = request.form['table2']
            if rel == "Many to One":
                fk = tab1 + tab2 + "mofk"
                fk = checkandcorrect(fk,tab1)
                ind = getPk(tab1)
                # <class 'tuple'>: (0, 'ФИО', 'text', 0, None, 1)
                st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk + " " + ind[2] + " REFERENCES " + tab2 + "(" + ind[1] + ")" + ";\n"
            elif rel == "One to One":
                fk1 = tab1 + tab2 + "oofk"
                fk2 = tab2 + tab1 + "oofk2"
                fk1 = checkandcorrect(fk1,tab1)
                fk2 = checkandcorrect(fk2,tab2)
                # <class 'tuple'>: (0, 'ФИО', 'text', 0, None, 1)
                st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk1 + " " + "INTEGER REFERENCES " + tab2 + "(" + fk2 + ")" +";\n"
                cursor.execute(st)
                st = "ALTER TABLE " + tab2 + " ADD COLUMN " + fk2 + " " + "INTEGER REFERENCES " + tab1 + "(" + fk1 + ")"+";\n"
            else:
                ind1 = getPk(tab1)
                ind2 = getPk(tab2)
                # st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk + " " + ind[2] + " REFERENCES " + tab2 + "(" + ind[1] + ")"
                nt = tab1 + "_" + tab2
                st = "CREATE TABLE " + nt
                st += " (" + ind1[1] + "1 " + ind1[2] + ","
                st += ind2[1] + "2 " + ind2[2] + ","
                st += "CONSTRAINT PKKEY PRIMARY KEY(" + ind1[1] + "1," + ind2[1] + "2),"
                st += "FOREIGN KEY(" + ind1[1] + "1) REFERENCES " + tab1 + "(" + ind1[1] + "),"
                st += "FOREIGN KEY(" + ind2[1] + "2) REFERENCES " + tab2 + "(" + ind2[1] + "))"+ ";\n"
            cursor.execute(st)
            st1 = "insert into tablerel (tab1, rel, tab2, mmtable) values ( '" + tab1 + "','" + rel + "','" + tab2 + "','" + nt + "' )"
            cursor.execute(st1)
            systemtables.append(nt)
            db.commit()
            f.write(st)
        except Exception as e:
            flash(e)
    f.close()
    rels = []
    cursor.execute("select * from tablerel")
    data = cursor.fetchall()
    for i in data:
        rels.append(i[1:])
    return render_template('reltables.html', tables=t, rels=rels)


@app.route('/table', methods=['POST', 'GET'])
def table():
    db = get_db()
    db.row_factory = make_dicts
    cursor = db.cursor()
    if request.method == 'POST':
        cursor.execute("select name from sqlite_master where type='table'")
        t = cursor.fetchall()
        tables = []
        for x in t:
            tables.append(x['name'])
        d = dict(request.form)
        tname = d['name'][0]
        if (tables != [] and tname in tables):
            flash("Table already exsits")
            return redirect('/tableEditor')
        try:
            mas = d['atr']
            prkey = d['prkey'][0]
            if (prkey in mas) and prkey != "auto":
                mas.remove(prkey)
            st = "create table " + tname + " ("
            header = 'insert into ' + tname + " ("
            w = ws[session['ind']]
            if prkey != 'auto':
                v = gettype(str(w.cell(row=2, column=int(prkey)).value))
                st += str(w.cell(row=1, column=int(prkey)).value) + " " + v + " primary key"
                header += str(w.cell(row=1, column=int(prkey)).value) + ","
            else:
                st += "id integer primary key autoincrement"
            for ww in mas:
                v = gettype(str(w.cell(row=2, column=int(ww)).value))
                st += ", " + str(w.cell(row=1, column=int(ww)).value) + " " + v
                header += str(w.cell(row=1, column=int(ww)).value) + ","
            st += ")"
            header = header[0:-1]
            header += ") values ("
            cursor.execute(st)
            db.commit()
        except Exception as e:
            flash("Failed to create table")
            return redirect('/tableEditor')
        f = open("static\script.txt", 'a')
        f.write(st + ";\n")
        c=0
        for i in range(2, w.max_row + 1):
            st = header
            if prkey != "auto":
                st += "'" + str(w.cell(row=i, column=int(prkey)).value) + "',"
            for j in mas:
                st += "'" + str(w.cell(row=i, column=int(j)).value) + "',"
            st = st[0:-1]
            st += ")"
            try:
                cursor.execute(st)
                f.write(st + ";\n")
            except Exception:
                c+=1
        f.close()
        db.commit()
        if c==0:
            flash("Successfully!")
        else:
            flash("Errors: "+str(c))
    else:
        tname = request.args.get('tab', '')  # Get
    data = None
    column = []
    try:
        cursor.execute("select * from " + tname)
        data = cursor.fetchall()
        for i in data[0]:
            column.append(i)
    except Exception as e:
        flash('No such table')
    return render_template('table.html', data=data, column=column)


@app.route('/tableEditor', methods=['POST', 'GET'])
def tableEditor():
    html_table = []
    ind= 0
    if request.method == 'POST':
        try:
            file = request.files['ex']
            global ws, l
            #str = file.filename.split('.')
            #if len(str) < 2 or str[len(str) - 1] != 'xlsx':
            workbook = load_workbook(file, data_only=True)
            ws = workbook.worksheets
            l = workbook.sheetnames
            workbook.close()
        except Exception:
            flash("Error: wrong file format.")
            return redirect('/')
    else:
        st = request.args.get('list', '')
        if l is None:
            flash("Choose excel file")
            return redirect('/')
        if st != '':
            for i in l:
                if i == st:
                    break
                ind += 1

    session["ind"] = ind
    for w in range(1, ws[ind].max_column + 1):
        html_table.append(ws[ind].cell(row=1, column=w).value)
    return render_template('tableEditor.html', html_table=html_table, list=l, thislist=l[session["ind"]])


@app.route('/')
def hello():
    return render_template('hello.html')


def myinit():
    st = """create table tablerel (
                id integer primary key autoincrement, 
                tab1 text not null, 
                rel text not null, 
                tab2 text not null,
                mmtable text
                )"""
    db = sqlite3.connect(pathdb)
    cursor = db.cursor()
    cursor.execute(st)
    db.commit()
    f = open("static\script.txt", 'w')
    f.close()

def getPort():
    from socket import socket,gethostbyname,AF_INET, SOCK_STREAM
    import random
    target = "localhost"
    targetIP = gethostbyname(target)
    s = socket(AF_INET, SOCK_STREAM)  # Инициализируем TCP-сокет
    while True:
        rid = random.randint(1000, 6500)  # Случайный порт
        result = s.connect_ex((targetIP,8080))
        if result != 0:
            s.close()
            return rid

if __name__ == "__main__":
    app.debug = True
    myinit()
    port = getPort()
    app.run(port =getPort())
    # host = "10.16.1.203", port =getPort()
