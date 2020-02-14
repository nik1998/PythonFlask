import uuid

from flask import Flask, request, session, g, redirect, url_for, render_template, flash, send_file
from werkzeug.datastructures import FileStorage
from openpyxl import load_workbook
import os
import sqlite3
from services import *

app = Flask(__name__)
# session - переменная в которой хранится информация о прошлых действиях пользователя
pathdb = os.path.join(os.path.dirname(__file__), 'table.db')
app.config.update(DATABASE=pathdb, DEBUG=True, SECRET_KEY='secretkey',
                  USERNAME='admin', PASSWORD='admin')

if os.path.exists(pathdb):
    os.remove(pathdb)


def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(pathdb)
    return db


def make_dicts(cursor, row):
    return dict((cursor.description[idx][0], value)
                for idx, value in enumerate(row))


@app.route('/downland', methods=['GET'])
def downland():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("select tab from userTables where user='" + session['id'] + "'")
    tab = cursor.fetchall()
    t = [i[0] for i in tab]
    with open('static\dump.sql', 'w') as f:
        for line in getTableDump(pathdb, ['Цена', 'Изделие', 'Производитель']):
            f.write('%s\n' % line)
    return send_file('static\dump.sql')


@app.route('/author', methods=['POST', 'GET'])
def author():
    return render_template("Author.html")


@app.route('/reltables', methods=['POST', 'GET'])
def reltables():
    db = get_db()
    cursor = db.cursor()
    st = ''
    if request.method == 'POST':
        try:
            tab1 = request.form['table1']
            rel = request.form['rel']
            tab2 = request.form['table2']
            if rel == "Many to One":
                fk = tab1 + tab2 + "mofk"
                fk = checkandcorrect(db, fk, tab1)
                ind = getPk(db, tab1)
                # <class 'tuple'>: (0, 'ФИО', 'text', 0, None, 1)
                st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk + " " + ind[2] + " REFERENCES " + tab2 + "(" + ind[
                    1] + ")" + ";\n"
            elif rel == "One to One":
                fk1 = tab1 + tab2 + "oofk"
                fk2 = tab2 + tab1 + "oofk2"
                ind1 = getPk(db, tab1)
                ind2 = getPk(db, tab2)
                fk1 = checkandcorrect(db, fk1, tab1)
                fk2 = checkandcorrect(db, fk2, tab2)
                # <class 'tuple'>: (0, 'ФИО', 'text', 0, None, 1)
                st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk1 + " " + ind2[2] + " REFERENCES " + tab2 + "(" + ind2[
                    1] + ")" + ";\n"
                cursor.execute(st)
                st = "ALTER TABLE " + tab2 + " ADD COLUMN " + fk2 + " " + ind1[2] + " REFERENCES " + tab1 + "(" + ind1[
                    1] + ")" + ";\n"
            else:
                ind1 = getPk(db, tab1)
                ind2 = getPk(db, tab2)
                nt = tab1 + "_" + tab2
                # st = "ALTER TABLE " + tab1 + " ADD COLUMN " + fk + " " + ind[2] + " REFERENCES " + tab2 + "(" + ind[1] + ")"
                st = "CREATE TABLE " + nt
                st += " (" + ind1[1] + "1 " + ind1[2] + ","
                st += ind2[1] + "2 " + ind2[2] + ","
                st += "CONSTRAINT PKKEY PRIMARY KEY(" + ind1[1] + "1," + ind2[1] + "2),"
                st += "FOREIGN KEY(" + ind1[1] + "1) REFERENCES " + tab1 + "(" + ind1[1] + "),"
                st += "FOREIGN KEY(" + ind2[1] + "2) REFERENCES " + tab2 + "(" + ind2[1] + "))" + ";\n"
            cursor.execute(st)
            if rel == "Many to Many":
                nt = tab1 + "_" + tab2
                st1 = "insert into userTables ( tab, user) values ('" + nt + "','" + session["id"] + "')"
                cursor.execute(st1)
            db.commit()
        except Exception as e:
            flash(str(e))
    cursor.execute("select tab from userTables where user='" + session['id'] + "'")
    tab = cursor.fetchall()
    return render_template('reltables.html', tables=tab)


@app.route('/editLines', methods=['POST'])
def editLines():
    try:
        db = get_db()
        db.row_factory = make_dicts
        tab = session['tab']
        cursor = db.cursor()
        cursor.execute("pragma table_info ('" + tab + "') ")
        data = cursor.fetchall()
        pr = ''
        d = dict(request.form)
        st = ' set'
        for txt in data:
            if txt['pk'] == 1:
                pr = txt['name']
            else:
                if request.form[txt['name']] != '':
                    st += ' ' + txt['name'] + " = '" + request.form[txt['name']] + "',"
                else:
                    st += ' ' + txt['name'] + ' = ' + '" "' + ','
        st = st[0:-1]
        ans = d[pr]
        str = "update " + session['tab'] + st + " where " + pr + " = " + ans
        cursor.execute(str)
        db.commit()
    except Exception as e:
        flash(str(e))
    return redirect(url_for('table', tab=session['tab']))


@app.route('/addLines', methods=['POST'])
def addLines():
    try:
        db = get_db()
        tab = session['tab']
        cursor = db.cursor()
        cursor.execute("pragma table_info ('" + tab + "') ")
        data = cursor.fetchall()
        pr = ''
        d = dict(request.form)
        st = ' set'
        for txt in data:
            if txt['pk'] == 1:
                pr = txt['name']
            else:
                if request.form[txt['name']] != '':
                    st += ' ' + txt['name'] + " = '" + request.form[txt['name']] + "',"
                else:
                    st += ' ' + txt['name'] + ' = ' + '" "' + ','
        st = st[0:-1]
        ans = d[pr]
        str = "update " + session['tab'] + st + " where " + pr + " = " + ans
        cursor.execute(str)
        db.commit()
    except Exception as e:
        flash(str(e))
    return redirect(url_for('table', tab=session['tab']))


@app.route('/table', methods=['POST', 'GET'])
def table():
    db = get_db()
    db.row_factory = make_dicts
    cursor = db.cursor()
    if request.method == 'POST':
        cursor.execute("select name from sqlite_master where type='table'")
        t = cursor.fetchall()
        tables = []
        for x in t:
            tables.append(x['name'])
        d = dict(request.form)
        tname = d['name']
        if (tables != [] and tname in tables):
            flash("Table already exsits")
            return redirect(url_for('tableEditor'))
        try:
            mas = []
            for i in range(0, len(d)):
                stt = "atr" + str(i + 1)
                if stt in d:
                    mas.append(d[stt])
            prkey = d['prkey']
            if (prkey in mas) and prkey != "auto":
                mas.remove(prkey)
            st = "create table " + tname + " ("
            header = 'insert into ' + tname + " ("
            ws = load_workbook(FileStorage(open(session['xls'], 'rb')), data_only=True).worksheets
            w = ws[session['ind']]
            if prkey != 'auto':
                v = gettype(str(w.cell(row=2, column=int(prkey)).value))
                st += str(w.cell(row=1, column=int(prkey)).value) + " " + v + " primary key"
                header += str(w.cell(row=1, column=int(prkey)).value) + ","
            else:
                st += "tabid integer primary key autoincrement"
            for ww in mas:
                v = gettype(str(w.cell(row=2, column=int(ww)).value))
                st += ", '" + str(w.cell(row=1, column=int(ww)).value) + "' " + v
                header += "'" + str(w.cell(row=1, column=int(ww)).value) + "',"
            st += ")"
            header = header[0:-1]
            header += ") values ("
            cursor.execute(st)
            st = "insert into userTables ( tab, user) values ('" + tname + "','" + session["id"] + "')"
            cursor.execute(st)
            db.commit()
        except Exception as e:
            flash(str(e))
            flash("Failed to create table")
            return redirect(url_for('tableEditor'))
        c = 0
        for i in range(2, w.max_row + 1):
            st = header
            if prkey != "auto":
                st += "'" + str(w.cell(row=i, column=int(prkey)).value) + "',"
            for j in mas:
                st += "'" + str(w.cell(row=i, column=int(j)).value) + "',"
            st = st[0:-1]
            st += ")"
            try:
                cursor.execute(st)
            except Exception:
                c += 1
        db.commit()
        if c == 0:
            flash("Successfully!")
        else:
            flash("Errors: " + str(c))
    else:
        tname = request.args.get('tab', '')
    session['tab'] = tname
    data = None
    column = []
    try:
        cursor.execute("select * from " + tname)
        data = cursor.fetchall()
        if data != []:
            for i in data[0]:
                column.append(i)
    except Exception as e:
        flash('No such table')
    return render_template('table.html', data=data, column=column)


@app.route('/tableEditor', methods=['POST', 'GET'])
def tableEditor():
    html_table = []
    ind = 0
    if request.method == 'POST':
        try:
            file = request.files['ex']
            # str = file.filename.split('.')
            # if len(str) < 2 or str[len(str) - 1] != 'xlsx':
            st = os.path.join(os.path.dirname(__file__), os.path.join('excelfiles', file.filename))
            file.save(st)
            session['xls'] = st
        except Exception:
            flash("Error: wrong file format.")
            return redirect(url_for('hello'))
    else:
        st = request.args.get('list', '')
        if 'xls' not in session:
            flash("Choose excel file")
            return redirect(url_for('hello'))
        if st != '':
            ind = int(st)
    f = FileStorage(open(session['xls'], 'rb'))
    workbook = load_workbook(f, data_only=True)
    workbook.close()
    session["ind"] = ind
    ws = workbook.worksheets
    l = workbook.sheetnames
    for w in range(1, ws[ind].max_column + 1):
        html_table.append(ws[ind].cell(row=1, column=w).value)
    return render_template('tableEditor.html', html_table=html_table, list=l, thislist=l[session["ind"]])


@app.route('/')
def hello():
    if 'id' not in session:
        session['id'] = str(uuid.uuid4())
    return render_template('hello.html')


if __name__ == "__main__":
    app.debug = True
    myinit(sqlite3.connect(pathdb))
    port = getPort()
    app.run(port=getPort())
    # host = "10.16.1.203", port =getPort()
